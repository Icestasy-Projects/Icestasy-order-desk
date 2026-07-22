import io
from datetime import date, datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, OutlineProperties

BRAND_FILL = PatternFill(start_color="FF8A3D", end_color="FF8A3D", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="21201D")
SUBTLE_FONT = Font(size=10, color="847E74", italic=True)
TOTAL_FONT = Font(bold=True)

STATUS_LABELS = {
    "draft": "Order Pending", "pending_approval": "In Progress", "approved": "In Progress",
    "in_production": "In Progress", "dispatched": "In Progress", "delivered": "Completed",
    "rejected": "Rejected", "cancelled": "Cancelled", "invoiced": "Invoiced",
}
PAY_LABELS = {
    "received": "Payment Completed", "pending": "Payment Pending", "failed": "Payment Failed",
    "refunded": "Refunded", "not_recorded": "Not Recorded",
}

COLUMNS = ["Order No", "Date", "Client", "Locality", "City", "Salesperson", "Status", "Payment Status", "Amount (INR)"]
COLUMN_WIDTHS = [16, 14, 28, 20, 16, 18, 16, 18, 14]

ORDER_LINE_COLUMNS = ["", "", "Flavour", "Format", "SKU Code", "Qty", "Unit Price", "Line Total", ""]
LINE_FILL = PatternFill(start_color="FFF3E8", end_color="FFF3E8", fill_type="solid")
LINE_FONT = Font(size=10, color="555555")
ORDER_HEADER_FILL = PatternFill(start_color="FFECD6", end_color="FFECD6", fill_type="solid")
ORDER_HEADER_FONT = Font(bold=True, size=11)

# Mumbai Metropolitan Region: reported as a single "Mumbai" group, matching the dashboard's By City tab.
MMR_CITIES = {"Mumbai", "Vasai-Virar", "Thane", "Navi Mumbai"}

REPORT_TYPE_LABELS = {
    "all": "All Orders", "month": "Month on Month", "city": "City-wise",
    "client": "Client-wise", "salesperson": "Salesperson-wise",
    "flavour": "Flavour-wise", "sku": "SKU-wise",
}
GROUP_LABELS = {"month": "Month", "city": "City", "client": "Client", "salesperson": "Salesperson"}

LINE_GROUP_KEY_FUNCS = {
    "flavour": lambda l: l.get("flavour_name") or "—",
    "sku": lambda l: l.get("sku_code") or "—",
}
LINE_GROUP_LABELS = {"flavour": "Flavour", "sku": "SKU Code"}
LINE_COLUMNS = ["Date", "Flavour", "SKU Code", "City", "Quantity", "Revenue (INR)"]
LINE_COLUMN_WIDTHS = [14, 30, 16, 16, 12, 16]


def _format_date(created_at: str) -> str:
    try:
        return datetime.fromisoformat(created_at.replace("Z", "+00:00")).strftime("%d %b %Y")
    except (ValueError, AttributeError):
        return created_at or ""


def _month_label(created_at: str) -> str:
    try:
        return datetime.fromisoformat(created_at.replace("Z", "+00:00")).strftime("%B %Y")
    except (ValueError, AttributeError):
        return created_at or "—"


def _report_city(o: dict) -> str:
    city = o.get("city")
    return "Mumbai" if city in MMR_CITIES else (city or "—")


GROUP_KEY_FUNCS = {
    "month": lambda o: _month_label(o.get("created_at", "")),
    "city": _report_city,
    "client": lambda o: o.get("client_name") or "—",
    "salesperson": lambda o: o.get("salesperson_name") or "—",
}


def _last_12_months_range() -> tuple:
    today = date.today()
    month_index = today.year * 12 + (today.month - 1) - 11
    start = date(month_index // 12, month_index % 12 + 1, 1)
    return start.isoformat(), today.isoformat()


def _in_range(created_at: str, date_from: str, date_to: str) -> bool:
    # Compare the "YYYY-MM-DD" prefix directly instead of parsing to a datetime —
    # created_at comes back from Postgres with varying offset formats (e.g. "+00"
    # vs "+00:00") depending on the client/runtime, which previously made
    # datetime.fromisoformat() silently reject rows on some Python versions.
    # Plain ISO-date-string comparison sorts identically to real dates and sidesteps
    # that entirely.
    date_str = (created_at or "")[:10]
    if not date_str:
        return True
    if date_from and date_str < date_from:
        return False
    if date_to and date_str > date_to:
        return False
    return True


def _sheet_header(ws, title_text: str, subtitle_text: str, last_col_letter: str):
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title_text
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = subtitle_text
    ws["A2"].font = SUBTLE_FONT


def _write_summary_sheet(ws, orders: list, report_type: str, subtitle_text: str):
    group_func = GROUP_KEY_FUNCS[report_type]
    group_label = GROUP_LABELS[report_type]
    columns = [group_label, "Orders", "Total Value (INR)"]
    last_col_letter = get_column_letter(len(columns))

    _sheet_header(ws, f"Icestasy Order Desk — {REPORT_TYPE_LABELS[report_type]} Report", subtitle_text, last_col_letter)

    header_row = 4
    for col, title in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = BRAND_FILL
        cell.alignment = Alignment(horizontal="center")

    totals = {}
    for o in orders:
        key = group_func(o) or "—"
        entry = totals.setdefault(key, {"count": 0, "value": 0.0})
        entry["count"] += 1
        entry["value"] += o.get("total_amount", 0) or 0

    row = header_row + 1
    grand_total = 0.0
    for name, entry in sorted(totals.items(), key=lambda kv: kv[1]["value"], reverse=True):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=entry["count"])
        amount_cell = ws.cell(row=row, column=3, value=entry["value"])
        amount_cell.number_format = "#,##0.00"
        grand_total += entry["value"]
        row += 1

    ws.cell(row=row, column=1, value="Total").font = TOTAL_FONT
    total_cell = ws.cell(row=row, column=3, value=grand_total)
    total_cell.font = TOTAL_FONT
    total_cell.number_format = "#,##0.00"

    for i, width in enumerate([28, 12, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"


def _write_orders_sheet(ws, orders: list, subtitle_text: str):
    last_col_letter = get_column_letter(len(COLUMNS))
    _sheet_header(ws, "Icestasy Order Desk — Orders Report", subtitle_text, last_col_letter)

    ws.sheet_properties.outlinePr = OutlineProperties(summaryBelow=False)

    header_row = 4
    for col, title in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = BRAND_FILL
        cell.alignment = Alignment(horizontal="center")

    amount_col = len(COLUMNS)
    row = header_row + 1
    total_value = 0.0
    for o in orders:
        values = [
            o.get("order_no"), _format_date(o.get("created_at", "")), o.get("client_name"),
            o.get("place"), o.get("city"), o.get("salesperson_name"),
            STATUS_LABELS.get(o.get("status"), o.get("status")),
            PAY_LABELS.get(o.get("payment_status"), o.get("payment_status")),
            o.get("total_amount", 0),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == amount_col:
                cell.number_format = "#,##0.00"
            cell.fill = ORDER_HEADER_FILL
            cell.font = ORDER_HEADER_FONT
        total_value += o.get("total_amount", 0) or 0
        row += 1

        lines = o.get("lines") or []
        if lines:
            for ln in lines:
                ws.cell(row=row, column=3, value=ln.get("flavour_name")).font = LINE_FONT
                ws.cell(row=row, column=4, value=ln.get("format_name")).font = LINE_FONT
                ws.cell(row=row, column=5, value=ln.get("sku_code")).font = LINE_FONT
                qty_cell = ws.cell(row=row, column=6, value=ln.get("quantity"))
                qty_cell.font = LINE_FONT
                qty_cell.number_format = "#,##0"
                up_cell = ws.cell(row=row, column=7, value=ln.get("unit_price"))
                up_cell.font = LINE_FONT
                up_cell.number_format = "#,##0.00"
                lt_cell = ws.cell(row=row, column=8, value=ln.get("line_total"))
                lt_cell.font = LINE_FONT
                lt_cell.number_format = "#,##0.00"
                for c in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row, column=c).fill = LINE_FILL
                ws.row_dimensions[row].outline_level = 1
                ws.row_dimensions[row].hidden = True
                row += 1

    total_label_col = amount_col - 1
    ws.cell(row=row, column=total_label_col, value="Total").font = TOTAL_FONT
    total_cell = ws.cell(row=row, column=amount_col, value=total_value)
    total_cell.font = TOTAL_FONT
    total_cell.number_format = "#,##0.00"

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = f"A{header_row + 1}"


def _write_flavour_summary_sheet(ws, lines: list, report_type: str, subtitle_text: str):
    group_func = LINE_GROUP_KEY_FUNCS[report_type]
    group_label = LINE_GROUP_LABELS[report_type]
    columns = [group_label, "Qty Sold", "Revenue (INR)"]
    last_col_letter = get_column_letter(len(columns))

    _sheet_header(ws, f"Icestasy Order Desk — {REPORT_TYPE_LABELS[report_type]} Report", subtitle_text, last_col_letter)

    header_row = 4
    for col, title in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = BRAND_FILL
        cell.alignment = Alignment(horizontal="center")

    totals = {}
    for l in lines:
        key = group_func(l) or "—"
        entry = totals.setdefault(key, {"qty": 0.0, "revenue": 0.0})
        entry["qty"] += l.get("quantity", 0) or 0
        entry["revenue"] += l.get("revenue", 0) or 0

    row = header_row + 1
    grand_qty = 0.0
    grand_revenue = 0.0
    for name, entry in sorted(totals.items(), key=lambda kv: kv[1]["qty"], reverse=True):
        ws.cell(row=row, column=1, value=name)
        qty_cell = ws.cell(row=row, column=2, value=entry["qty"])
        qty_cell.number_format = "#,##0"
        rev_cell = ws.cell(row=row, column=3, value=entry["revenue"])
        rev_cell.number_format = "#,##0.00"
        grand_qty += entry["qty"]
        grand_revenue += entry["revenue"]
        row += 1

    ws.cell(row=row, column=1, value="Total").font = TOTAL_FONT
    qty_total_cell = ws.cell(row=row, column=2, value=grand_qty)
    qty_total_cell.font = TOTAL_FONT
    qty_total_cell.number_format = "#,##0"
    rev_total_cell = ws.cell(row=row, column=3, value=grand_revenue)
    rev_total_cell.font = TOTAL_FONT
    rev_total_cell.number_format = "#,##0.00"

    for i, width in enumerate([30, 12, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"


def _write_flavour_detail_sheet(ws, lines: list, subtitle_text: str):
    last_col_letter = get_column_letter(len(LINE_COLUMNS))
    _sheet_header(ws, "Icestasy Order Desk — Flavour Sales Detail", subtitle_text, last_col_letter)

    header_row = 4
    for col, title in enumerate(LINE_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = BRAND_FILL
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    total_qty = 0.0
    total_revenue = 0.0
    for l in sorted(lines, key=lambda x: x.get("created_at") or ""):
        values = [
            _format_date(l.get("created_at", "")), l.get("flavour_name"), l.get("sku_code"),
            l.get("city"), l.get("quantity", 0), l.get("revenue", 0),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == 5:
                cell.number_format = "#,##0"
            if col == 6:
                cell.number_format = "#,##0.00"
        total_qty += l.get("quantity", 0) or 0
        total_revenue += l.get("revenue", 0) or 0
        row += 1

    ws.cell(row=row, column=4, value="Total").font = TOTAL_FONT
    qty_cell = ws.cell(row=row, column=5, value=total_qty)
    qty_cell.font = TOTAL_FONT
    qty_cell.number_format = "#,##0"
    rev_cell = ws.cell(row=row, column=6, value=total_revenue)
    rev_cell.font = TOTAL_FONT
    rev_cell.number_format = "#,##0.00"

    for i, width in enumerate(LINE_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"


def build_flavour_sales_workbook(lines: list, role_label: str, full_name: str,
                                  report_type: str = "flavour", date_from: str = None, date_to: str = None,
                                  orders_without_lines: int = 0, value_without_lines: float = 0.0) -> io.BytesIO:
    if report_type not in ("flavour", "sku"):
        report_type = "flavour"
    if date_from or date_to:
        lines = [l for l in lines if _in_range(l.get("created_at", ""), date_from, date_to)]

    generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
    range_text = f" · {date_from or '…'} to {date_to or '…'}" if (date_from or date_to) else ""
    total_qty = sum(l.get("quantity", 0) or 0 for l in lines)
    subtitle_text = f"{full_name} · {role_label} · Generated {generated} · {total_qty:.0f} units sold{range_text}"
    if orders_without_lines:
        subtitle_text += (f" · Pre-tax product revenue; excludes {orders_without_lines} orders "
                           f"(₹{value_without_lines:,.2f}) with no recorded line items")

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_flavour_summary_sheet(summary_ws, lines, report_type, subtitle_text)
    detail_ws = wb.create_sheet("Detail")
    _write_flavour_detail_sheet(detail_ws, lines, subtitle_text)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


CLIENT_SUMMARY_COLUMNS = ["Client", "City", "Orders", "Total Amount (INR)", "Amount Paid (INR)", "Amount Pending (INR)"]
CLIENT_SUMMARY_WIDTHS = [30, 16, 10, 20, 20, 20]


def _write_client_summary_sheet(ws, orders: list, subtitle_text: str):
    columns = CLIENT_SUMMARY_COLUMNS
    last_col_letter = get_column_letter(len(columns))
    _sheet_header(ws, "Icestasy Order Desk — Client Summary", subtitle_text, last_col_letter)

    header_row = 4
    for col, title in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = BRAND_FILL
        cell.alignment = Alignment(horizontal="center")

    clients = {}
    for o in orders:
        name = o.get("client_name") or "—"
        entry = clients.setdefault(name, {"city": o.get("city") or "—", "count": 0, "total": 0.0, "paid": 0.0})
        entry["count"] += 1
        entry["total"] += o.get("total_amount", 0) or 0
        entry["paid"] += o.get("amount_paid", 0) or 0

    row = header_row + 1
    grand_total = grand_paid = grand_pending = 0.0
    grand_count = 0
    for name, e in sorted(clients.items(), key=lambda kv: kv[1]["total"], reverse=True):
        pending = e["total"] - e["paid"]
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=e["city"])
        ws.cell(row=row, column=3, value=e["count"])
        for col, val in [(4, e["total"]), (5, e["paid"]), (6, pending)]:
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = "#,##0.00"
        grand_count += e["count"]
        grand_total += e["total"]
        grand_paid += e["paid"]
        grand_pending += pending
        row += 1

    ws.cell(row=row, column=1, value="Total").font = TOTAL_FONT
    ws.cell(row=row, column=3, value=grand_count).font = TOTAL_FONT
    for col, val in [(4, grand_total), (5, grand_paid), (6, grand_pending)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = TOTAL_FONT
        c.number_format = "#,##0.00"

    for i, width in enumerate(CLIENT_SUMMARY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"


def build_orders_workbook(orders: list, role_label: str, full_name: str,
                           report_type: str = "all", date_from: str = None, date_to: str = None) -> io.BytesIO:
    if report_type not in REPORT_TYPE_LABELS:
        report_type = "all"
    if report_type == "month" and not date_from and not date_to:
        date_from, date_to = _last_12_months_range()
    if date_from or date_to:
        orders = [o for o in orders if _in_range(o.get("created_at", ""), date_from, date_to)]

    generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
    range_text = f" · {date_from or '…'} to {date_to or '…'}" if (date_from or date_to) else ""
    subtitle_text = f"{full_name} · {role_label} · Generated {generated} · {len(orders)} orders{range_text}"

    wb = Workbook()
    if report_type == "all":
        ws = wb.active
        ws.title = "Orders"
        _write_orders_sheet(ws, orders, subtitle_text)
    else:
        summary_ws = wb.active
        summary_ws.title = "Summary"
        _write_summary_sheet(summary_ws, orders, report_type, subtitle_text)
        detail_ws = wb.create_sheet("Orders")
        _write_orders_sheet(detail_ws, orders, subtitle_text)

    client_ws = wb.create_sheet("Client Summary")
    _write_client_summary_sheet(client_ws, orders, subtitle_text)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
